import pandas as pd
from datetime import datetime
import numpy as np
from typing import List, Dict, Tuple
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

class BankReconciliation:
    def __init__(self):
        self.df1 = None  # Premier fichier Excel
        self.df2 = None  # Deuxième fichier Excel
        self.results = []  # Résultats du rapprochement
        self.green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        self.red_fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')

    def load_files(self, file1_path: str, file2_path: str) -> None:
        """Charge les deux fichiers Excel"""
        try:
            # Charger les fichiers complets
            df1_temp = pd.read_excel(file1_path)
            df2_temp = pd.read_excel(file2_path)
            
            # Sélectionner les colonnes avec leurs vrais noms
            self.df1 = pd.DataFrame({
                'Date': df1_temp.iloc[:, 0],        # Colonne A
                'Description': df1_temp.iloc[:, 1],  # Colonne B
                'MontantC': pd.to_numeric(df1_temp.iloc[:, 2], errors='coerce'),    # Colonne C (Crédit)
                'MontantD': pd.to_numeric(df1_temp.iloc[:, 3], errors='coerce'),    # Colonne D (Débit)
            })
            
            self.df2 = pd.DataFrame({
                'Date': df2_temp.iloc[:, 1],        # Colonne B
                'Description': df2_temp.iloc[:, 6],  # Colonne G
                'MontantL': pd.to_numeric(df2_temp.iloc[:, 11], errors='coerce'),   # Colonne L (Débit)
                'MontantM': pd.to_numeric(df2_temp.iloc[:, 12], errors='coerce'),   # Colonne M (Crédit)
            })
            
            # Remplacer les NaN par None pour une meilleure gestion
            self.df1 = self.df1.replace({np.nan: None})
            self.df2 = self.df2.replace({np.nan: None})
            
        except Exception as e:
            raise ValueError(f"Impossible de charger les fichiers. Erreur: {str(e)}")

    def clean_description(self, desc: str) -> str:
        """Nettoie la description en suivant les mêmes règles que le VBA"""
        if not isinstance(desc, str):
            return ""
        
        desc = desc.upper().strip()
        prefixes = ['DE:', 'POUR:', 'REMISE:', 'VIRT.', 'MME', 'MR']
        for prefix in prefixes:
            desc = desc.replace(prefix, '')
        return desc.strip()

    def match_descriptions(self, desc1: str, desc2: str) -> bool:
        """Compare deux descriptions selon les règles du VBA"""
        # Extraire les mots de chaque description
        words1 = [w for w in self.clean_description(desc1).split() if w]
        words2 = [w for w in self.clean_description(desc2).split() if w]
        
        if not words1 or not words2:
            return False
        
        # Compter les correspondances
        matches = sum(1 for w1 in words1 if w1 in words2)
        
        # Retourner vrai si au moins la moitié des mots correspondent
        return matches >= len(words1)/2 or matches >= len(words2)/2

    def reconcile(self) -> pd.DataFrame:
        """Effectue le rapprochement bancaire selon les règles du VBA"""
        results = []
        matched_rows2 = set()
        matched_rows1 = set()
        special_amounts = {}  # Pour suivre les montants spéciaux (C et M)

        # Premier passage : correspondance D-L par montant uniquement
        for idx1, row1 in self.df1.iterrows():
            if pd.notna(row1['MontantD']) and row1['MontantD'] is not None:
                for idx2, row2 in self.df2.iterrows():
                    if idx2 not in matched_rows2 and pd.notna(row2['MontantL']) and row2['MontantL'] is not None:
                        try:
                            # Comparaison des montants en valeur absolue
                            if abs(round(float(row1['MontantD']), 2)) == abs(round(float(row2['MontantL']), 2)):
                                results.append({
                                    'Date_1': row1['Date'],
                                    'Description_1': row1['Description'],
                                    'Description_2': row2['Description'],
                                    'Montant_1': row1['MontantD'],
                                    'Montant_2': row2['MontantL'],
                                    'Status': 'Matched',
                                    'Type': 'D-L'
                                })
                                matched_rows2.add(idx2)
                                matched_rows1.add(idx1)
                                break
                        except (ValueError, TypeError):
                            continue

        # Deuxième passage : correspondance C-M par montant uniquement
        for idx1, row1 in self.df1.iterrows():
            if idx1 not in matched_rows1 and pd.notna(row1['MontantC']):
                for idx2, row2 in self.df2.iterrows():
                    if idx2 not in matched_rows2 and pd.notna(row2['MontantM']):
                        try:
                            if abs(round(float(row1['MontantC']), 2)) == abs(round(float(row2['MontantM']), 2)):
                                key = str(abs(round(float(row1['MontantC']), 2)))
                                if key not in special_amounts:
                                    results.append({
                                        'Date_1': row1['Date'],
                                        'Description_1': row1['Description'],
                                        'Description_2': row2['Description'],
                                        'Montant_1': row1['MontantC'],
                                        'Montant_2': row2['MontantM'],
                                        'Status': 'Matched',
                                        'Type': 'C-M'
                                    })
                                    matched_rows2.add(idx2)
                                    matched_rows1.add(idx1)
                                    special_amounts[key] = True
                                    break
                        except (ValueError, TypeError):
                            continue

        # Ajouter les lignes non appariées de df1
        for idx1, row1 in self.df1.iterrows():
            if idx1 not in matched_rows1:
                if pd.notna(row1['MontantD']) or pd.notna(row1['MontantC']):
                    results.append({
                        'Date_1': row1['Date'],
                        'Description_1': row1['Description'],
                        'Description_2': 'Non trouvé',
                        'Montant_1': row1['MontantD'] if pd.notna(row1['MontantD']) else row1['MontantC'],
                        'Montant_2': None,
                        'Status': 'Unmatched',
                        'Type': 'D' if pd.notna(row1['MontantD']) else 'C'
                    })

        return pd.DataFrame(results)

    def _create_match_entry(self, row1, row2, match_type):
        """Crée une entrée pour une correspondance trouvée"""
        return {
            'Date_1': row1['Date'],
            'Description_1': row1['Description'],
            'Description_2': row2['Description'],
            'Montant_1': row1['MontantD'] if match_type == 'D-L' else row1['MontantC'],
            'Montant_2': row2['MontantL'] if match_type == 'D-L' else row2['MontantM'],
            'Status': 'Matched',
            'Type': match_type
        }

    def _add_unmatched_rows(self, results, matched_rows1, matched_rows2):
        """Ajoute les lignes non appariées au résultat"""
        # Lignes non appariées de Feuil1
        for idx1, row1 in self.df1.iterrows():
            if idx1 not in matched_rows1:
                results.append({
                    'Date_1': row1['Date'],
                    'Description_1': row1['Description'],
                    'Description_2': 'Non trouvé',
                    'Montant_1': row1['MontantD'] if pd.notna(row1['MontantD']) else row1['MontantC'],
                    'Montant_2': None,
                    'Status': 'Unmatched',
                    'Type': 'File1'
                })

        # Lignes non appariées de Feuil2
        for idx2, row2 in self.df2.iterrows():
            if idx2 not in matched_rows2:
                results.append({
                    'Date_1': row2['Date'],
                    'Description_1': 'Non trouvé',
                    'Description_2': row2['Description'],
                    'Montant_1': None,
                    'Montant_2': row2['MontantL'] if pd.notna(row2['MontantL']) else row2['MontantM'],
                    'Status': 'Unmatched',
                    'Type': 'File2'
                })

    def export_to_excel(self, output_path: str) -> None:
        """Exporte les résultats dans un fichier Excel"""
        df_results = self.reconcile()
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapprochement"

        # Écrire les en-têtes
        headers = ['Date', 'Description 1', 'Description 2', 'Montant 1', 'Montant 2', 'Statut', 'Type']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Écrire les données et appliquer le formatage
        for row_idx, row in df_results.iterrows():
            ws_row = row_idx + 2
            ws.cell(row=ws_row, column=1, value=row['Date_1'])
            ws.cell(row=ws_row, column=2, value=row['Description_1'])
            ws.cell(row=ws_row, column=3, value=row['Description_2'])
            ws.cell(row=ws_row, column=4, value=row['Montant_1'])
            ws.cell(row=ws_row, column=5, value=row['Montant_2'])
            ws.cell(row=ws_row, column=6, value=row['Status'])
            ws.cell(row=ws_row, column=7, value=row['Type'])

            # Appliquer la couleur selon le statut
            fill = self.green_fill if row['Status'] == 'Matched' else self.red_fill
            for col in range(1, 8):
                ws.cell(row=ws_row, column=col).fill = fill

        # Sauvegarder le fichier
        wb.save(output_path)

class BankReconciliationGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Rapprochement Bancaire")
        self.root.geometry("800x600")
        
        self.reconciliation = BankReconciliation()
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.output_path = tk.StringVar()
        
        self.create_menu()
        self.create_widgets()
        
    def create_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # Menu Fichier
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Fichier", menu=file_menu)
        file_menu.add_command(label="Nouveau rapprochement", command=self.reset_form)
        file_menu.add_separator()
        file_menu.add_command(label="Quitter", command=self.root.quit)
        
        # Menu Aide
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Aide", menu=help_menu)
        help_menu.add_command(label="À propos", command=self.show_about)
        
    def create_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Sélection des fichiers
        ttk.Label(main_frame, text="Fichier 1 (Relevé bancaire):").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.file1_path, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(main_frame, text="Parcourir", command=lambda: self.browse_file(self.file1_path)).grid(row=0, column=2)
        
        ttk.Label(main_frame, text="Fichier 2 (Comptabilité):").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.file2_path, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(main_frame, text="Parcourir", command=lambda: self.browse_file(self.file2_path)).grid(row=1, column=2)
        
        ttk.Label(main_frame, text="Fichier de sortie:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.output_path, width=50).grid(row=2, column=1, padx=5)
        ttk.Button(main_frame, text="Parcourir", command=self.browse_output_file).grid(row=2, column=2)
        
        # Boutons d'action
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=3, column=0, columnspan=3, pady=20)
        
        ttk.Button(buttons_frame, text="Lancer le rapprochement", command=self.run_reconciliation).pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="Réinitialiser", command=self.reset_form).pack(side=tk.LEFT, padx=5)
        
        # Zone de log
        self.log_text = tk.Text(main_frame, height=15, width=70)
        self.log_text.grid(row=4, column=0, columnspan=3, pady=10)
        
        # Barre de progression
        self.progress = ttk.Progressbar(main_frame, length=300, mode='determinate')
        self.progress.grid(row=5, column=0, columnspan=3, pady=10)
        
    def browse_file(self, path_var):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            path_var.set(filename)
            
    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_path.set(filename)
            
    def reset_form(self):
        self.file1_path.set("")
        self.file2_path.set("")
        self.output_path.set("")
        self.log_text.delete(1.0, tk.END)
        self.progress['value'] = 0
        
    def show_about(self):
        messagebox.showinfo(
            "À propos",
            "Application de Rapprochement Bancaire\nVersion 1.0\n\nCette application permet de rapprocher deux fichiers Excel contenant des transactions bancaires."
        )
        
    def log_message(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def run_reconciliation(self):
        if not all([self.file1_path.get(), self.file2_path.get(), self.output_path.get()]):
            messagebox.showerror("Erreur", "Veuillez sélectionner tous les fichiers nécessaires.")
            return
            
        try:
            self.log_message("Début du rapprochement...")
            self.progress['value'] = 20
            
            # Chargement des fichiers
            self.reconciliation.load_files(self.file1_path.get(), self.file2_path.get())
            self.log_message("Fichiers chargés avec succès")
            self.progress['value'] = 50
            
            # Effectuer le rapprochement et exporter
            self.reconciliation.export_to_excel(self.output_path.get())
            self.progress['value'] = 100
            
            self.log_message("Rapprochement terminé avec succès!")
            messagebox.showinfo("Succès", "Le rapprochement a été effectué avec succès!")
            
        except Exception as e:
            self.log_message(f"Erreur: {str(e)}")
            messagebox.showerror("Erreur", f"Une erreur est survenue:\n{str(e)}")
            
def main():
    root = tk.Tk()
    app = BankReconciliationGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main() 